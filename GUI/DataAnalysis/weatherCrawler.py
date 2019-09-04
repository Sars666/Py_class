import codecs
import requests, os
from bs4 import BeautifulSoup
from lxml import etree
import re
from openpyxl import Workbook, load_workbook
import time


# 获取url
def urlAnalysis(url, encoding='gbk'):
    req = requests.get(url)
    req.encoding = 'gb18030'
    reqState = req.status_code
    return req, reqState


# 用BS4保存Html文件
def saveHtml(req, filename):
    soup = BeautifulSoup(req.text, features="lxml")
    with codecs.open(filename + ".html", "w", "utf-8") as f:
        f.write(soup.prettify())
    return soup


# 用etree得到地址和标题
def getHref(req):
    xmlContent = etree.HTML(req.text)
    cityHrefList = xmlContent.xpath('//*[@id="content"]/div[3]/dl/dd/a/@href')
    cityNameList = xmlContent.xpath('//*[@id="content"]/div[3]/dl/dd/a/text()')
    # print(cityHrefList)
    # print(cityNameList)
    return cityHrefList, cityNameList


# 创建空文件夹
def createEmptyDir(targetDir):
    if not os.path.isdir(targetDir):
        os.mkdir(targetDir)


# 创建Excel
def createNewExcel(excelName):
    try:
        load_workbook(excelName)
    except FileNotFoundError:
        wb = Workbook()
        wb.save(excelName)


# 存储数据
def saveData(ws, monthData, dateTimeList, titleList):
    columns = 7
    rows = int(len(monthData) / columns)
    for i in range(4):
        ws.cell(row=1, column=1 + i).value = titleList[i]
    for r in range(rows):
        ws.cell(row=r + 1 + 1, column=1).value = dateTimeList[r]
        for c in range(columns):
            ws.cell(row=r + 1 + 1, column=c + 1 + 1).value = monthData[r * columns + c]


# 保存excel并命名
def saveDataIntoExcel(dateTimeList, month, cityName, dateTime, yearDataList, filename):
    titleList = ['日期', '天气状况', '气温', '风力风向']
    targetDir = os.path.join(os.path.dirname(os.path.abspath(__file__)), filename)
    createEmptyDir(targetDir)
    excelName = filename + '/' + cityName.strip() + '.xlsx'
    createNewExcel(excelName)
    wb = load_workbook(excelName)
    # process
    if dateTime not in wb.sheetnames:
        ws = wb.create_sheet(title=dateTime, index=month)
    else:
        ws = wb[dateTime]
    if ws.cell(row=3, column=3).value != None:
        print(dateTime + ' already exist')
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
        return
    else:
        saveData(ws, yearDataList[month], dateTimeList, titleList)
        print(dateTime + ' downloaded')
        print(time.strftime("%Y-%m-%d %H:%M:%S", time.localtime()))
    wb.save(excelName)


# 从爬下来的数据中筛选出有用数据
def informationFilter(list):
    informationList = []
    for id in range(len(list)):
        who = re.findall('.*[\S]', list[id])
        if who != []:
            for i in range(len(who)):
                who[i] = who[i].strip()
                informationList.append(who[i])
    return informationList


# 由地址下载信息
def downloadInfo(mainurl, hrefList, nameList, maxCityNum, filename, encoding):
    if hrefList != []:
        for cityNum in range(len(hrefList)):
            while maxCityNum > 0:
                cityUrl = mainurl + hrefList[0]  # 选取对应参数可爬其他城市数据,这里只要北京
                print(mainurl + hrefList[cityNum])
                cityReq = urlAnalysis(cityUrl, encoding)[0]  # 同上
                print(nameList[cityNum])
                cityXmlContent = etree.HTML(cityReq.text)
                timeHref = cityXmlContent.xpath('//*[@id="content"]/div/ul/li/a/@href')  # 201101-201905
                timeName = cityXmlContent.xpath('//*[@id="content"]/div/ul/li/a/text()')
                print(timeHref)
                print(timeName)
                yearDataList = []
                for time in range(len(timeHref)):
                    monthNum = (timeHref[time])[-11:-5]
                    currentTimeUrl = mainurl + (f
                    '/lishi/beijing/month/{monthNum}.html')
                    currentTimeReq = requests.get(currentTimeUrl)
                    currentTimeXmlContent = etree.HTML(currentTimeReq.text)
                    dateTimeList = currentTimeXmlContent.xpath('//*[@id="content"]/table/tr/td/a/text()')
                    dateTimeList = informationFilter(dateTimeList)
                    timeContentList = currentTimeXmlContent.xpath('//*[@id="content"]/table/tr/td/text()')
                    yearDataList.append(informationFilter(timeContentList))
                    saveDataIntoExcel(dateTimeList, month=time, cityName=nameList[cityNum], dateTime=timeName[time],
                                      yearDataList=yearDataList, filename=filename)
                print(yearDataList)
                maxCityNum -= 1
                wb = load_workbook(filename + '/' + nameList[cityNum].strip() + '.xlsx')
                wb.remove(wb["Sheet"])
                wb.save(filename + '/' + nameList[cityNum].strip() + '.xlsx')

    else:
        print('Failed')


# 主程序
def main(mainurl, reqState, req, maxCityNum, filename, encoding):
    if reqState == 200:
        print('Url Request Accepted')
        # saveHtml(req, filename)
        hrefList, nameList = getHref(req)
        downloadInfo(mainurl, hrefList, nameList, maxCityNum, filename, encoding)
    else:
        print('Url Request Denied')


'''
mainurl = 'http://www.tianqihoubao.com'
url = 'http://www.tianqihoubao.com/lishi'
encoding = 'gbk'
head = {'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_6)'
                      ' AppleWebKit/605.1.15 (KHTML, like Gecko)'
                      ' Version/12.0.1 Safari/605.1.15',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',}
filename = 'WeatherHistory'
req, reqState = urlAnalysis(url,filename)
maxSetNumber = 1

if reqState == 200:
    main(mainurl, reqState, req, maxSetNumber, filename,encoding)
else:
    print(reqState)

'''
