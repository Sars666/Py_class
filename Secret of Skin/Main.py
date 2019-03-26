import jieba
import matplotlib.pyplot as plt
from math import log10
import regex as re
from collections import Counter
from openpyxl import load_workbook
from openpyxl.styles import Font,Alignment


#读取已有工作簿
file_name = 'statistics.xlsx'
wb = load_workbook(file_name)

#创建新工作表
st_name = "result"
if st_name in wb.sheetnames:
    wb.remove(wb[st_name])
    ws0 = wb.create_sheet(title=st_name)
else:
    ws0 = wb.create_sheet(title=st_name)

#读取文本文件
filename = 'SecretOfSkin.txt'
with open(filename, 'rt') as f:
    book = f.read()

#分出汉字characterList
characterList = []
for line in book:
    for character in line:
        if '\u4E00' <= character <= '\u9FA5':
            characterList.append(character)
#结巴库分词
seg_list = jieba.lcut(book)
wordList = "/ ".join(seg_list).split('/ ')

#词频统计wordDic
def get_wordDic(List):
    wordDic = {}
    wordBook = List
    for word in wordBook:
        wordDic[word] = wordDic.get(word, 0) + 1
    for k in wordDic.keys():
        if not '\u4E00' <= k <= '\u9FA5':
            wordDic[k] = 0
    return wordDic

#结果排序
def result_sort(wordDic):
    result = [(v, k) for k, v in wordDic.items()]
    result.sort(reverse= True)
    result = [(v, k) for k, v in result]
    return result

#打印结果表
def input_result(result):
    global col
    rank = 1
    for j in range(len(result)):
        if result[j][1]!= 0:
            for i in range(len(result[j])):
                ws.cell(row = j+2, column = col+i).value = str(result[j][i])
                ws.cell(row = j+2, column = col+i).font = font
                ws.cell(row = j+2, column = col+i).alignment = alignment
            ws.cell(row = j+2, column = col+2).value = rank
            ws.cell(row = j+2, column = col+2).font = font
            ws.cell(row = j+2, column = col+2).alignment = alignment
            rank += 1
    col += 4


#去除汉字外的字符,并写入横纵坐标列表
def get_rankAndcount(result):
    rank = []
    count = []
    for i in result:
        if i[1] != 0:
            count.append(log10(i[1]))
    for i in range(len(count)) :
        rank.append(log10(i+1))
    return (count,rank)

#设置excel标题行
def set_title():
    value = ['word','count','rank']
    for i in range(0,3):
        for j in range(1,4):
            ws.cell(row = 1, column = 4*i+j).value = value[j-1]
            ws.cell(row = 1, column = 4*i+j).font = title_font
            ws.cell(row = 1, column = 4*i+j).alignment = title_alignment

#全局变量初始化
col = 1

wordDic_1 = get_wordDic(wordList)
wordDic_2 = get_wordDic(characterList)
result_1 = result_sort(wordDic_1)
result_2 = result_sort(wordDic_2)
result_3 = []

rank_1 = get_rankAndcount(result_1)[1]
count_1 = get_rankAndcount(result_1)[0]
rank_2 = get_rankAndcount(result_2)[1]
count_2 = get_rankAndcount(result_2)[0]

#matplotlib画图
plt.plot(rank_1, count_1, marker='o', mec='red', mfc='w',label="word Frequency")
plt.plot(rank_2, count_2, marker='+', mec='blue', mfc='w',label="character Frequency")
plt.xlabel("rank")
plt.xlim(-0.5,4.5)
plt.xticks(range(5),('0','$10^1$','$10^2$','$10^3$','$10^4$'))
plt.yticks(range(5),('0','$10^1$','$10^2$','$10^3$','$10^4$'))
plt.ylim(-0.5,4.5)
plt.ylabel("count")
plt.title("Zipf's law")
plt.grid(True)
plt.legend()
#plt.show()
ws = wb[st_name]

match_regex = '皮肤([^“”，。、：\s/《》]{2,3})[“”，。、：\s/《》]'
anything = []
for i in re.finditer(match_regex, book):
    who = i.group(1)
    anything.append(who)
c = Counter(anything)
for k,v in c.most_common(20):
    result_3.append((k,v))

#结果转到excel
font = Font(size=16)
alignment=Alignment(horizontal='center')
title_font = Font(size=20, bold=True)
title_alignment = alignment

set_title()
input_result(result_1)
input_result(result_2)
input_result(result_3)
#保存工作簿
print(ws.cell)
wb.save(file_name)
print('结果在excel文件中显示')