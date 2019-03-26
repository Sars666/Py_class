from openpyxl import load_workbook

#读取已有工作簿
file_name = 'statistics.xlsx'
wb = load_workbook(file_name)

#创建新工作表
st_name_0 = "正向标准归一化结果"
if st_name_0 in wb.sheetnames:
    wb.remove(wb[st_name_0])
    ws0 = wb.create_sheet(title=st_name_0)
else:
    ws0 = wb.create_sheet(title=st_name_0)

#结果输入
ws = wb[st_name_0]
ws.cell(row = j, column = 1).value = 1
#保存工作簿
wb.save(file_name)
print('ok')